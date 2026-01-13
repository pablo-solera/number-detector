import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


class ExcelExporter:
    def __init__(self, output_path):
        self.output_path = output_path

    def export(self, rows):
        if not rows:
            print("\n⚠ No hay datos para exportar")
            return

        # Crear DataFrame
        df = pd.DataFrame(
            rows,
            columns=["Archivo", "Número", "Motor"]
        )

        # Agrupar por archivo: mostrar nombre solo en primera fila
        df_grouped = df.copy()

        # Identificar dónde cambia el archivo
        df_grouped['archivo_changed'] = df_grouped['Archivo'] != df_grouped['Archivo'].shift(1)

        # Donde NO haya cambio de archivo, dejar la celda vacía
        df_grouped.loc[~df_grouped['archivo_changed'], 'Archivo'] = ''

        # Eliminar columna auxiliar
        df_grouped = df_grouped.drop('archivo_changed', axis=1)

        # Exportar a Excel
        df_grouped.to_excel(self.output_path, index=False, sheet_name='Números Rojos')

        # Aplicar formato profesional
        self._apply_formatting()

        print(f"\n✓ Resultados exportados a: {self.output_path}")
        print(f"  Total de registros: {len(rows)}")
        print(f"  Archivos procesados: {df['Archivo'].nunique()}")

    def _apply_formatting(self):
        """
        Aplica formato profesional al Excel generado
        """
        try:
            wb = load_workbook(self.output_path)
            ws = wb.active

            # Estilos para encabezados
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)

            # Estilos para bordes
            thin_border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )

            # Formatear encabezados (primera fila)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border

            # Formatear datos (resto de filas)
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='left', vertical='center')

            # Ajustar ancho de columnas
            ws.column_dimensions['A'].width = 18  # Archivo
            ws.column_dimensions['B'].width = 12  # Número
            ws.column_dimensions['C'].width = 20  # Motor

            # Congelar primera fila (encabezados)
            ws.freeze_panes = 'A2'

            # Guardar cambios
            wb.save(self.output_path)

        except Exception as e:
            print(f"⚠ Advertencia: No se pudo aplicar formato al Excel: {e}")